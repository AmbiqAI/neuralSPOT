# save as xlsx_audit.py and run:  python xlsx_audit.py your_file.xlsx
import sys
from openpyxl import load_workbook

def audit(path):
    wb = load_workbook(path, data_only=False, read_only=False)
    print("Sheets:", wb.sheetnames)
    try:
        names = [n.name for n in wb.defined_names.definedName]
        print("Defined names:", names)
    except Exception:
        pass
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    print("Header count =", len(headers))
    print("Headers (first 12):", headers[:12])
    # Tables
    tbls = getattr(ws, "tables", None)
    tbls = list(tbls.values()) if isinstance(tbls, dict) else list(getattr(ws, "_tables", []))
    for t in tbls:
        try:
            print("Table:", t.displayName, "ref:", t.ref)
        except Exception:
            print("Table present (name/ref unreadable)")
    # First few #REF!
    ref_hits = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        for c in row:
            v = c.value
            if isinstance(v, str) and "#REF!" in v or getattr(c, "data_type", None) == "e":
                ref_hits.append((c.coordinate, v))
                if len(ref_hits) >= 15:
                    break
        if len(ref_hits) >= 15:
            break
    if ref_hits:
        print("First #REF! occurrences:")
        for coord, txt in ref_hits:
            print(" ", coord, ":", txt)
    else:
        print("No #REF! strings detected.")
    # print first non-header row in the Derived sheet
    # print as a list with 2 columns header and value
    ws = wb["Config"]
    headers = [c.value for c in ws[1]]
    # print("Headers (first 12):", headers[:12])
    for row in ws.iter_rows(min_row=1, values_only=True):
        print([(headers[i], c) for i, c in enumerate(row)])
        break
    wb.close()

if __name__ == "__main__":
    audit(sys.argv[1])
