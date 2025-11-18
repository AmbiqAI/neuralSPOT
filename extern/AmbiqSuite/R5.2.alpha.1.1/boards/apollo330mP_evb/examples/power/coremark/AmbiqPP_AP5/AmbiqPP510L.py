# This is a Ambiq Power Profiling tool source code
#
# Copyright (c) 2022, Ambiq Micro
# NOTE: This GUI is for Apollo510L debugging
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are not permitted
# Rev 2.0 Added new file scan function and highlights the differences between steps

from cgi import test
from pickle import GLOBAL
from unittest import TestProgram
import serial
import json
import datetime
import time
import pandas as pd
import openpyxl
import shutil
import subprocess
import os
import configparser
import sys
import signal

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import formatting, styles
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, CellIsRule, FormulaRule


from subprocess import call

bFirstTime = True
outputpath = './Outputs/'

plotPwrVal= []
plotPwrVal= [0 for i in range(37)]
plotCLKVal= [0 for i in range(37)]
timervalArr = [0 for i in range(16)]

#this is the raw timer reg reading
timerRawVal = [0 for i in range(16)]

HFRC = "96M"
LFRC = "1.024K"
TURBO = "192M"
XT = "32.768K"

HFRCVal = 96000
XTVal = 32.768
LFRCVal = 1.024
TMRPINB = 32.768
TURBOVal = 192000

def readConfig():
    # Check whether the specified path exists or not
    isExist = os.path.exists(outputpath)
    if not isExist:
        os.makedirs(outputpath)
        print("The output folder is created!")
    else:
        #clear the old one first
        shutil.rmtree(outputpath)
        os.makedirs(outputpath)
        print("The output folder is created!")


    config = configparser.ConfigParser()

    config.read('AmbiqPPConfig.ini')
    portconfig = config['DEFAULT']['COMPORT']

    connection = serial.Serial(port="COM%s"%portconfig, baudrate=115200, timeout = 2)
    connection.reset_input_buffer()
    connection.flushInput()
    connection.flushOutput()
    time.sleep(0.1)

    print("Ready to capture!")
    return connection

#Save the file to designated folder
def saveFile(xlsworkbook, filename):
    completeName = os.path.join(outputpath, filename)
    xlsworkbook.save(filename=completeName)

def extract_element_from_json(obj, path):
    '''
    Extracts an element from a nested dictionary or
    a list of nested dictionaries along a specified path.
    If the input is a dictionary, a list is returned.
    If the input is a list of dictionary, a list of lists is returned.
    obj - list or dict - input dictionary or list of dictionaries
    path - list - list of strings that form the path to the desired element
    '''
    def extract(obj, path, ind, arr):
        '''
            Extracts an element from a nested dictionary
            along a specified path and returns a list.
            obj - dict - input dictionary
            path - list - list of strings that form the JSON path
            ind - int - starting index
            arr - list - output list
        '''
        key = path[ind]
        if ind + 1 < len(path):
            if isinstance(obj, dict):
                if key in obj.keys():
                    extract(obj.get(key), path, ind + 1, arr)
                else:
                    arr.append(None)
            elif isinstance(obj, list):
                if not obj:
                    arr.append(None)
                else:
                    for item in obj:
                        extract(item, path, ind, arr)
            else:
                arr.append(None)
        if ind + 1 == len(path):
            if isinstance(obj, list):
                if not obj:
                    arr.append(None)
                else:
                    for item in obj:
                        arr.append(item.get(key, None))
            elif isinstance(obj, dict):
                arr.append(obj.get(key, None))
            else:
                arr.append(None)
        return arr
    if isinstance(obj, dict):
        return extract(obj, path, 0, [])
    elif isinstance(obj, list):
        outer_arr = []
        for item in obj:
            outer_arr.append(extract(item, path, 0, []))
        return outer_arr


#extract power information source and write to the data file using PowerTemplate.xlsx
def extract_power_and_clk(SnapN, pwrctrl, clkctrl, mcuctrl, timerctrl):
    MCUPERFREQ = int(extract_element_from_json(jObj, ["PWRCTRL", "MCUPERFREQ"])[0])
    MEMPWRSTATUS = int(extract_element_from_json(jObj, ["PWRCTRL", "MEMPWRSTATUS"])[0])
    SSRAMPWRST = int(extract_element_from_json(jObj, ["PWRCTRL", "SSRAMPWRST"])[0])
    #DSP0PERFREQ = int(extract_element_from_json(jObj, ["PWRCTRL", "DSP0PERFREQ"])[0])
    #DSP1PERFREQ = int(extract_element_from_json(jObj, ["PWRCTRL", "DSP1PERFREQ"])[0])
    VRSTATUS = int(extract_element_from_json(jObj, ["PWRCTRL", "VRSTATUS"])[0])
    AUDSSPWRSTATUS = int(extract_element_from_json(jObj, ["PWRCTRL", "AUDSSPWRSTATUS"])[0])
    #AUDADCSTATUS = int(extract_element_from_json(jObj, ["PWRCTRL", "AUDADCSTATUS"])[0])
    ADCSTATUS =  int(extract_element_from_json(jObj, ["PWRCTRL", "ADCSTATUS"])[0])
    DEVPWRSTATUS = int(extract_element_from_json(jObj, ["PWRCTRL", "DEVPWRSTATUS"])[0])

    #Now MCUctrl
    PDMCTRL =  int(extract_element_from_json(jObj, ["MCUCTRL", "PDMCTRL"])[0])

    #Now clock
    #Notice that we need to check GLOBEN first to make sure the controls are enabled
    GLOBEN = int(extract_element_from_json(jObj, ["Timers", "GLOBEN"])[0])
    CLOCKENSTAT = int(extract_element_from_json(jObj, ["CLK", "CLOCKENSTAT"])[0])
    CLOCKEN2STAT = int(extract_element_from_json(jObj, ["CLK", "CLOCKEN2STAT"])[0])
    CLOCKEN3STAT = int(extract_element_from_json(jObj, ["CLK", "CLOCKEN3STAT"])[0])
    LFRCCTRL = int(extract_element_from_json(jObj, ["CLK", "LFRCCTRL"])[0])
    CLKOUT = int(extract_element_from_json(jObj, ["CLK", "CLKOUT"])[0])

    #And Timers, which associates with CLOCK info
    #this is the raw timer reg reading
    for i in range (0,16):
        timerRawVal[i]  = int(extract_element_from_json(jObj, ["Timers", "CTRL%d"%i])[0])

    #Lastly, GPIO
    GPIO0 =  int(extract_element_from_json(jObj, ["PWRCTRL", "FPIOEN0"])[0])
    GPIO1 =  int(extract_element_from_json(jObj, ["PWRCTRL", "FPIOEN1"])[0])
    GPIO2 =  int(extract_element_from_json(jObj, ["PWRCTRL", "FPIOEN2"])[0])

    #Now save the JSON fields in xls table for PP tools to use
    # Start by opening the spreadsheet and selecting the main sheet
    workbookp = load_workbook(filename="PowerTemplate.xlsx")
    sheet = workbookp.active

    #CPU and memory
    regval = MCUPERFREQ & (0x3)
    clkstr = "24M" if regval==0 else "96M" if regval == 1 else "192M"
    sheet["E2"] = clkstr

    #DTCM
    regval = MEMPWRSTATUS & 0x8
    sheet["C3"] = sheet["D3"] = "N"  if regval ==0 else "Y"
    sheet["C3"] = "8K"  if regval ==1 else "128K" if regval == 2 else "384K"
    sheet["E3"] = clkstr

    #NVM0, Cache 0, 2
    for i in range (3, 6):
        regval = MEMPWRSTATUS & (0x1<<i)
        sheet["C%d"%(i+1)] = sheet["D%d"%(i+1)] = "Y" if regval else "N"
        sheet["E%d"%(i+1)] = clkstr

    #SRAM
    regval = SSRAMPWRST & 0x3
    sheet["C7"] = sheet["D7"] = "N" if regval == 0 or regval == 2 else "Y"
    sheet["E7"] = clkstr if regval == 1 or regval == 3 else ""
    sheet["C8"] = sheet["D8"] = "N" if regval == 0 or regval == 1 else "Y"
    sheet["E8"] = clkstr if regval == 2 or regval == 3 else ""

    #DSP
    #regval = DSP0PERFREQ & 0x1
    #sheet["E9"] = "48M" if regval==0 else "192M" if regval == 1 else "384M"
    #regval = DSP1PERFREQ & 0x1
    #sheet["E10"] = "48M" if regval==0 else "192M" if regval == 1 else "384M"
    sheet["E9"] = "48M"
    sheet["E10"] = "48M"

    #SIMOBUCK, MEMLDO, CORELDO
    regval = VRSTATUS & (0x3 << 4)
    sheet["C11"] = sheet["D11"] = "N"  if regval ==0 else "LP" if regval == 0x2 else "ACT"
    regval = VRSTATUS & (0x3 << 2)
    sheet["C12"] = sheet["D12"] = "N"  if regval ==0 else "LP" if regval == 0x2 else "ACT"
    regval = VRSTATUS & (0x3)
    sheet["C13"] = sheet["D13"] = "N"  if regval ==0 else "LP" if regval == 0x2 else "ACT"

    # #AUDADC
    # regval = AUDSSPWRSTATUS & (0x1 << 10)
    # sheet["C14"] = sheet["D14"] = "N"  if regval ==0 else "Y"
    # sheet["E14"] = clkstr if regval else ""

    # regval = AUDADCSTATUS
    # tempstr = ""
    # tempstr += "BandGap"  if regval &(0x1<<1) == 0 else ""
    # tempstr += "TempSensor"  if regval &(0x1<<2) == 0 else ""
    # tempstr += "ADCVbatt"  if regval &(0x1<<3) == 0 else ""
    # tempstr += "RefKeep"  if regval &(0x1<<4) == 0 else ""
    # tempstr += "RefBuff"  if regval &(0x1<<5) == 0 else ""
    # sheet["C15"] = sheet["D15"] = tempstr if tempstr !="" else "N"
    # sheet["E15"] = clkstr if tempstr !="" else ""

    regval = ADCSTATUS
    tempstr = ""
    tempstr += "BandGap"  if regval &(0x1<<1) == 0 else ""
    tempstr += "TempSensor"  if regval &(0x1<<2) == 0 else ""
    tempstr += "ADCVbatt"  if regval &(0x1<<3) == 0 else ""
    tempstr += "RefKeep"  if regval &(0x1<<4) == 0 else ""
    tempstr += "RefBuff"  if regval &(0x1<<5) == 0 else ""
    sheet["C16"] = sheet["D16"] = tempstr if tempstr !="" else "N"
    sheet["E16"] = clkstr if tempstr !="" else ""

    #Peripherals
    for i in range (0, 27):
        regval = DEVPWRSTATUS & (0x1<<i)
        sheet["C%d"%(i+17)] = sheet["D%d"%(i+17)] = "Y" if regval else "N"
        sheet["E%d"%(i+17)] = clkstr if regval else ""

    #PDM
    regval = PDMCTRL & (0x1<<i)
    sheet["C44"] = sheet["D44"] = "Y" if regval else "N"
    sheet["E44"] = clkstr if regval else ""

    #GPIOs
    tempstr = ""
    for c in range (0,3):
        regval = GPIO0 if c==0 else GPIO1 if c==1 else GPIO2
        for i in range (0, 32):
            tempstr += "%d,"%((i+1) + c*32) if (regval & (0x1 << i)) else ""
    if (tempstr == ""):
        sheet["C57"] = sheet["D57"] = "N"

    ######################## CLK ######################################
    bCheckTime = True
    for i in range (0,32):
        regval = CLOCKENSTAT & (0x1 <<i)
        if (i == 0):
            sheet["D30"] = "Y" if regval else "N"
        elif (i == 13 and regval == 0):
            for c in range(45, 53):
                sheet["D%d"%c] = "N"
                bCheckTime = False
        elif (i >= 14 and i <=29 and bCheckTime):
            if (regval == 1 and (GLOBEN & 0x1 << (i-14) or GLOBEN & 0x1 << (i-13))):
                sheet["D%d"%(54+(i-14)/2)] = "Y"
                if (i-14) %2 == 0:
                    sheet["E%d"%(54+(i-14)/2)] = extract_timer_freq(i-14, (timerRawVal[i] & (0xFF << 8)) >> 8)
            else:
                sheet["D%d"%(54+(i-14)/2)] = "N"

    ############# LFRC, HFRC, RTC and CLKOUT ##########################
    regval = LFRCCTRL & 0x3
    if (regval == 0):
        sheet["D54"] = "Y"
        sheet["E54"] = LFRC
    else:
        sheet["D54"] = "N"

    regval = CLOCKEN3STAT & 0x2000000
    if (regval != 0):
        sheet["D53"] = "Y"
        sheet["E53"] = HFRC
    else:
        sheet["D53"] = "N"

    regval = CLOCKEN3STAT & 0x10000000
    sheet["D55"] = "Y"
    sheet["E55"] = XT

    regval = CLKOUT & (0x1 <<7)
    if (regval !=0 ):
        sheet["D56"] = "Y"
        sheet["E56"] = extract_clkout_freq(CLKOUT & 0x1F)
    else:
        sheet["D56"] = "N"

    ###################### Write to the File #############################

    filename = "Power_%d_%02d%02d_%02d%02d%02d.xlsx" %(SnapN, now.month, now.day, now.hour, now.minute, now.second)
    saveFile(workbookp, filename)
    workbookp.close()
    return

#extract CLKOUT source and return the results
def extract_clkout_freq(CLKOUT):
    if (CLKOUT == 0x0):
        clkoutval = LFRCVal
    elif (0x1 <= CLKOUT and CLKOUT >= 0x5):
        clkoutval = XTVal / pow(2, CLKOUT)
    elif (CLKOUT == 0x10):
        clkoutval = 1000
    elif (CLKOUT == 0x16):
        clkoutval = XTVal/pow(2,21)
    elif (CLKOUT == 0x17):
        clkoutval = XTVal
    elif (CLKOUT == 0x18):
        clkoutval = 100
    elif (CLKOUT == 0x19 or CLKOUT == 0x22 or CLKOUT == 0x32):
        clkoutval = HFRCVal
    elif (CLKOUT == 0x1A or CLKOUT == 0x1B or CLKOUT == 0x1C):
        clkoutval = XTVal/pow(2, (CLKOUT-0x1A + 2))
    elif (CLKOUT >= 0x1D or CLKOUT <= 0x120):
        clkoutval = XTVal/pow(2, (CLKOUT-0x1D + 6))
    elif (CLKOUT >= 0x1D or CLKOUT <= 0x120):
        clkoutval = XTVal/pow(2, (CLKOUT-0x1D + 6))
    elif (CLKOUT == 0x23):
        clkoutval = LFRCVal/2
    elif (CLKOUT == 0x24):
        clkoutval = LFRCVal/32
    elif (CLKOUT == 0x25):
        clkoutval = LFRCVal/512
    elif (CLKOUT == 0x26):
        clkoutval = LFRCVal/32768
    elif (CLKOUT == 0x27):
        clkoutval = XTVal/256
    elif (CLKOUT == 0x28):
        clkoutval = XTVal/8192
    elif (CLKOUT == 0x29):
        clkoutval = XTVal/pow(2, 16)
    elif (CLKOUT == 0x28):
        clkoutval = XTVal/8192
    elif (CLKOUT == 0x2A):
        clkoutval = LFRCVal/16
    elif (CLKOUT == 0x2B):
        clkoutval = LFRCVal/128
    elif (CLKOUT == 0x2C):
        clkoutval = LFRCVal/1024
    elif (CLKOUT == 0x2D):
        clkoutval = LFRCVal/4096
    elif (CLKOUT == 0x2E):
        clkoutval = LFRCVal/pow(2,20)
    elif (CLKOUT == 0x2F):
        clkoutval = HFRCVal/pow(2,16)
    elif (CLKOUT == 0x30):
        clkoutval = HFRCVal/pow(2,24)
    elif (CLKOUT == 0x31):
        clkoutval = LFRCVal/16
    elif (CLKOUT == 0x33):
        clkoutval = HFRCVal/8
    elif (CLKOUT == 0x35):
        clkoutval = XTVal
    elif (CLKOUT == 0x36):
        clkoutval = XTVal/16
    elif (CLKOUT == 0x37):
        clkoutval = LFRCVal/32
    elif (CLKOUT == 0x39):
        clkoutval = LFRCVal
    elif (CLKOUT == 0x3A):
        clkoutval = 6000000
    elif (CLKOUT == 0x3B or CLKOUT == 0x3C):
        clkoutval = 24000000

    return clkoutval

#extract timer clock source and return the results
def extract_timer_freq(timern, timenval):
    if (timenval == 0x0):
        timerval = HFRCVal /4
    elif (timenval == 0x1):
        timerval = HFRCVal /16
    elif (timenval == 0x2):
        timerval = HFRCVal /64
    elif (timenval == 0x3):
        timerval = HFRCVal /256
    elif (timenval == 0x4):
        timerval = HFRCVal /1024
    elif (timenval == 0x5):
        timerval = HFRCVal /4096
    elif (timenval == 0x6):
        timerval = LFRCVal
    elif (timenval == 0x7):
        timerval = LFRCVal /2
    elif (timenval == 0x8):
        timerval = LFRCVal /32
    elif (timenval == 0x9):
        timerval = LFRCVal /1024
    elif (timenval == 0xA):
        timerval = XTVal
    elif (timenval == 0xB):
        timerval = XTVal /2
    elif (timenval == 0xC):
        timerval = XTVal /4
    elif (timenval == 0xD):
        timerval = XTVal /8
    elif (timenval == 0xE):
        timerval = XTVal /16
    elif (timenval == 0xF):
        timerval = XTVal /32
    elif (timenval == 0x10):
        timerval = XTVal /128
    elif (timenval == 0x11):
        timerval = 100

 ########### Be careful of the order here, we need to assign the clock value first before assign to other timers
    return timerval

print("Ambiq Power Profiling Tool V1.0 Starting ...")
print("Please wait for ready signal before reset your board ...")
connection = readConfig()
bCaptureFileReady = False
bStartParse = False

while True:

    data = connection.readline().decode('ascii', errors='ignore')

    if (bCaptureFileReady == True):
        delta = int(datetime.datetime.now().second - lastnow.second)
        print("Waiting for all snapshots to be captured...")
        if (delta >= 5):
            bStartParse = True
            bCaptureFileReady = False


    if data.startswith("{\"PWRCTRL\""):
        f = open(outputpath + 'capture.json','a')
        data=str(data)
        f.write(data)
        f.close()
        lastnow = datetime.datetime.now()
        print("JSON captured at:", lastnow)
        bCaptureFileReady = True

    #This routine is to capture our snapshot from JSON
    if bStartParse:
    #if data.startswith("{\"PWRCTRL\""):
        bStartParse = False
        f = open(outputpath + 'capture.json','r')
        lines = f.readlines()

        for line in lines:
            print("Start to parsing captured JSON...")
            # print(line)
            jObj = json.loads(line)
            #json_str = json.dumps(data, skipkeys = True,  allow_nan = True, indent=4)
            #print(jObj)

            pwrctrl = jObj['PWRCTRL']
            mcuctrl = jObj['MCUCTRL']
            sysctrl = jObj['SYSCTRL']
            clkctrl = jObj['CLK']
            stimerctrl = jObj['STIMER']
            timerctrl = jObj['TIMER']
            pdmctrl = jObj['PDM']


            snapNumber = pwrctrl["SnapN"]
            now = datetime.datetime.now()
            print("Parsing JSON for snapshot %d" %snapNumber)

            #Parse the information for power and clock
            # extract_power_and_clk(snapNumber, pwrctrl, clkctrl, mcuctrl, timerctrl)

            #Now save the capture under 2 cases: 1. No sequenced files 2. Sequence file found
            snapRange= int(snapNumber/10)*10
            target = "Snapshot_%d_%02d%02d"%(snapRange, now.month, now.day)
            xlsfilename = "xlsx"
            write2default = True

            files = os.listdir(outputpath)
            for file in files:
                if (file.endswith(xlsfilename)):
                    if file.startswith(target):
                        write2default = False
                        existingfile = file
                        print(existingfile)
                        break

            if (write2default == False):
                filename = outputpath + existingfile
                workbook = load_workbook(filename)
                sheet = workbook.active
                print("writing to an existing file...")
            else:
                workbook = load_workbook(filename="AP510LTemplate.xlsx")
                sheet = workbook.active
                print("writing to a new file...")


            # JSON fields in xls table for PP tools to use
            # Start by opening the spreadsheet and selecting the main sheet
            # workbook = load_workbook(filename="AP4Template.xlsx")

            redFill = PatternFill(start_color='FFFF0000',
                            end_color='FFFF0000',
                            fill_type='solid')

            pwrctrl_start_row = 5
            pwrctrl_end_row = 54

            clk_start_row = 56
            clk_end_row = 67

            sys_start_row = 69
            sys_end_row = 89

            stimer_start_row = 91
            stimer_end_row = 115

            timer_start_row = 117
            timer_end_row = 220

            mcuctrl_start_row = 222
            mcuctrl_end_row = 308

            pdmctrl_start_row = 310
            pdmctrl_end_row = sheet.max_row

            if (write2default == True):
                sheet["D2"] = True if (pwrctrl["Singleshot"] == 1) else False
                snapNumber = sheet["D3"] = pwrctrl["SnapN"]

                #Power ctrl first
                for i in range (pwrctrl_start_row, pwrctrl_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(pwrctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        #sheet["A%d"%i].fill = redFill
                        #sheet["C%d"%i].fill = redFill
                        sheet["D%d"%i].fill = redFill
                    #print (sheet["D%d"%i].value)

                #Now check the clk ctrl
                for i in range (clk_start_row, clk_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(clkctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        sheet["D%d"%i].fill = redFill

                #Now check the sys ctrl
                for i in range (sys_start_row, sys_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(sysctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        sheet["D%d"%i].fill = redFill

                #Now check the stimers
                for i in range (stimer_start_row, stimer_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(stimerctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        sheet["D%d"%i].fill = redFill

                #Now check the timers
                for i in range (timer_start_row, timer_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(timerctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        sheet["D%d"%i].fill = redFill

                #Now check the mcu ctrl
                for i in range (mcuctrl_start_row, mcuctrl_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(mcuctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        sheet["D%d"%i].fill = redFill

                #Last, check the pdm ctrl or specific block internal registers
                for i in range (pdmctrl_start_row, pdmctrl_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["D%d"%i] = hex(pdmctrl[index])
                    if (sheet["C%d"%i].value != sheet["D%d"%i].value):
                        sheet["D%d"%i].fill = redFill

                # Save the spreadsheet
                filename="Snapshot_%d_%02d%02d_%02d%02d%02d.xlsx" %(snapNumber, now.month, now.day, now.hour, now.minute, now.second)
                saveFile(workbook, filename)
                workbook.close()

            #writes to an existing file
            else:
                ws = workbook['Sheet1']
                new_col = ws.max_column
                prev_col = ws.max_column - 1

                newColName = chr(ord('A')+new_col)
                prvColName = chr(ord('A')+new_col - 1)

                sheet["%s2"%newColName] = True if (pwrctrl["Singleshot"] == 1) else False
                snapNumber = sheet["%s3"%newColName] = pwrctrl["SnapN"]

                #Power ctrl first
                for i in range (pwrctrl_start_row, pwrctrl_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(pwrctrl[index])
                    if (sheet["%s%d"%(newColName,i)].value != sheet["%s%d"%(prvColName,i)].value):
                        #sheet["%s%d"%(prvColName,i)].fill = redFill
                        sheet["%s%d"%(newColName,i)].fill = redFill
                    #print (sheet["D%d"%i].value)

                #Now check the clk ctrl
                for i in range (clk_start_row, clk_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(clkctrl[index])
                    if (sheet["%s%d"%(prvColName,i)].value != sheet["%s%d"%(newColName,i)].value):
                        #sheet["%s%d"%(prvColName,i)].fill = redFill
                        sheet["%s%d"%(newColName,i)].fill = redFill

                #Now check the sys clk ctrl
                for i in range (sys_start_row, sys_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(sysctrl[index])
                    if (sheet["%s%d"%(prvColName,i)].value != sheet["%s%d"%(newColName,i)].value):
                        #sheet["%s%d"%(prvColName,i)].fill = redFill
                        sheet["%s%d"%(newColName,i)].fill = redFill

                #Now check the stimers
                for i in range (stimer_start_row, stimer_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(stimerctrl[index])
                    if (sheet["%s%d"%(prvColName,i)].value != sheet["%s%d"%(newColName,i)].value):
                        #sheet["%s%d"%(prvColName,i)].fill = redFill
                        sheet["%s%d"%(newColName,i)].fill = redFill

                #Now check the timers
                for i in range (timer_start_row, timer_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(timerctrl[index])
                    if (sheet["%s%d"%(prvColName,i)].value != sheet["%s%d"%(newColName,i)].value):
                        #sheet["%s%d"%(prvColName,i)].fill = redFill
                        sheet["%s%d"%(newColName,i)].fill = redFill

                #Now check the mcu ctrl
                for i in range (mcuctrl_start_row, mcuctrl_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(mcuctrl[index])
                    if (sheet["%s%d"%(prvColName,i)].value != sheet["%s%d"%(newColName,i)].value):
                        #sheet["%s%d"%(prvColName,i)].fill = redFill
                        sheet["%s%d"%(newColName,i)].fill = redFill

                #Lastly, check the pdm ctrl or specific block internal registers
                for i in range (pdmctrl_start_row, pdmctrl_end_row + 1):
                    index = str(sheet["A%d"%i].value)
                    sheet["%s%d"%(newColName,i)] = hex(pdmctrl[index])
                    if (sheet["%s%d"%(prvColName,i)].value != sheet["%s%d"%(newColName,i)].value):
                        sheet["%s%d"%(newColName,i)].fill = redFill

                # Save the spreadsheet
                saveFile(workbook, file)
                workbook.close()


        ######################## End of Power Domain ############################

        ######################## CLK Domain #####################################
        #CPU HFRC clock must be enabled and output is enabled

        #Timers, 8 timers with A and B options
        # TimerBlockEnable = CLOCKENSTAT & 0x8000
        ######################## End of CLK Domain ##############################

        #Now do the conditional formatting
        # if (bFirstTime):
        #     bFirstTime = False
        #     #call(["python", "AmbiqPPGUI.py"])
        #     subprocess.Popen(["python", "AmbiqPP5GUI.py"])

    #This routine is capturing normal user debugging information if needed
    else:
        print(data)





